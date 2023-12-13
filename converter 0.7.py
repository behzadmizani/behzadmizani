import csv
import os
import openpyxl
from openpyxl.styles import PatternFill, Font , Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
import tkinter
from tkinter import filedialog
from tkinter import *



root = tkinter.Tk()

input_filez = filedialog.askopenfilenames(parent=root, title='Select Resualt Files as TXT')
template_file = filedialog.askopenfilename(parent=root, title='Select Template File as XLSX')
#output_file = 'C:/Users/mizani/Documents/audit proj/beta 0.92/output.xlsx'
output_file = filedialog.asksaveasfilename(parent=root, title='Save Output File as XLSX',defaultextension=".xlsx")

# Open xlsx template
wb = openpyxl.load_workbook(template_file)
sheet = wb['Red Hat 8.0']

# Color picker
yellow = "00FFFF00"
red = "FF0000"
green = "00B050"
orange = "FD9203"
gray = "BFBFBF"

# Initialize column counter
column = 'F'



#customize file name columns 2
fname_fill = PatternFill(start_color=gray, end_color=gray, fill_type="solid")
fname_font = Font(bold=True, size=12)
fname_align = Alignment(horizontal=CENTER,vertical=CENTER)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


# Iterate through multiple input files
for input_file in input_filez:
    # Open each txt file
    with open(input_file, 'r') as f:
        lines = f.readlines()
        #for next column 
        column = chr(ord(column) + 1)
        #file name add to first column row
        file_name = "Status \n"
        file_name += os.path.basename(input_file)  # Extract filename from path
        file_name = (os.path.splitext(file_name)[0])
        #file_name = file_name[34:71]
        xname  = file_name.split()
        statusname = "Status \n"
        #extract word 7(ip address) & word 8(host name)
        statusname += xname[7]
        statusname += "\n"
        statusname += xname[8]
        filename_cell = sheet[column + str(2)] #column set columns and str(set num of row) for write file name (host and ip)
       #add file name to selected column 
        filename_cell.value = statusname

    #customize font and fill for first column
        filename_cell.fill = fname_fill
        filename_cell.font = fname_font
        filename_cell.alignment = fname_align
        filename_cell.border = thin_border
        

    # Iterate through each line in the TXT file
    for i, line in enumerate(lines):
        status = line.split(',')[-1].strip()

        # Extract status value
        result_cell = sheet[column + str(i + 3)]  # Add 'i + 3' to increment column after each file
        result_cell.value = status

        # Determine fill color based on status
        if status == 'Fail':
            cell_fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
        elif status == 'Pass':
            cell_fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        elif status == 'N/A':
            cell_fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
        else:
            cell_fill = PatternFill(start_color=orange, end_color=orange, fill_type="solid")

        # Apply fill to the result cell
        result_cell.fill = cell_fill
        result_cell.font = fname_font
        result_cell.alignment = fname_align
        result_cell.border = thin_border
        

# Save the xlsx template
wb.save(output_file)

print("*** Conversion is Complete ***")
